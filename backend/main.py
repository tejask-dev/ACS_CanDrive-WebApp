"""
ACS Can Drive Backend API
==========================

This is the main FastAPI application for the ACS Can Drive web application.
It provides RESTful API endpoints for managing can drive events, students, 
donations, leaderboards, and street reservations.

Key Features:
- Student and teacher management
- Donation tracking and processing
- Real-time leaderboard calculations
- Street reservation system with Google Maps integration
- Admin authentication and authorization
- CSV import/export functionality
- Daily leaderboard reset functionality

Architecture:
- FastAPI framework for REST API
- SQLAlchemy ORM for database operations
- SQLite database for data persistence
- JWT-based authentication for admin access

Author: ACS Can Drive Development Team
Created: 2025
"""

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from database import init_db, get_db
from routers import admin, events, students, donations, map_reservations
from config import CURRENT_EVENT_ID
import os

# Initialize FastAPI application
app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for now
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

init_db()

# Simple database helper for SQLite
def get_db_simple():
    """Get database connection - simple for SQLite"""
    from database import SessionLocal
    return SessionLocal()

# Database context manager to ensure connections are closed
def get_db_with_context():
    """Get database connection with automatic cleanup"""
    from database import SessionLocal
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Ensure admin user exists on startup
def ensure_admin_user():
    from database import SessionLocal
    from models import Admin
    from datetime import datetime
    import hashlib
    try:
        db = SessionLocal()
        existing_admin = db.query(Admin).filter(Admin.username == 'ACS_CanDrive').first()
        if not existing_admin:
            admin = Admin(
                username='ACS_CanDrive',
                password_hash=hashlib.sha256('Assumption_raiders'.encode()).hexdigest(),
                created_at=datetime.now()
            )
            db.add(admin)
            db.commit()
            print('✅ Admin user created on startup!')
        else:
            print('✅ Admin user already exists')
        db.close()
    except Exception as e:
        print(f'❌ Error ensuring admin user: {e}')

ensure_admin_user()

# Ensure event exists on startup
def ensure_event():
    from database import SessionLocal
    from models import Event
    from datetime import datetime
    try:
        db = SessionLocal()
        existing_event = db.query(Event).filter(Event.id == CURRENT_EVENT_ID).first()
        if not existing_event:
            event = Event(
                id=CURRENT_EVENT_ID,
                name=f'ACS Can Drive {datetime.now().year}',
                start_date=datetime.now(),
                created_at=datetime.now()
            )
            db.add(event)
            db.commit()
            print(f'✅ Event {CURRENT_EVENT_ID} created successfully')
        else:
            print(f'✅ Event {CURRENT_EVENT_ID} already exists')
        db.close()
    except Exception as e:
        print(f'❌ Error ensuring event: {e}')

ensure_event()

app.include_router(admin.router, prefix="/api/auth", tags=["auth"])
# app.include_router(events.router, prefix="/api/events", tags=["events"])  # Temporarily disabled due to route conflict with direct leaderboard endpoint
# app.include_router(students.router, prefix="/api/events/{event_id}/students", tags=["students"])  # Temporarily disabled due to route conflict
# app.include_router(donations.router, prefix="/api/events/{event_id}/donations", tags=["donations"])  # Temporarily disabled due to route conflict
# app.include_router(map_reservations.router, prefix="/api/events/{event_id}/map-reservations", tags=["map"])  # Temporarily disabled due to route conflict

@app.get("/")
def read_root():
    return {"msg": "ACS Can Drive API running"}

@app.get("/favicon.ico")
def favicon():
    from fastapi.responses import Response
    return Response(status_code=204)

@app.get("/health")
def health_check():
    """Health check endpoint to monitor database connectivity"""
    try:
        db = get_db_simple()
        
        # Test basic table access
        from models import Student, Teacher, Donation, MapReservation
        student_count = db.query(Student).count()
        teacher_count = db.query(Teacher).count()
        donation_count = db.query(Donation).count()
        reservation_count = db.query(MapReservation).count()
        
        db.close()
        return {
            "status": "healthy", 
            "database": "connected", 
            "msg": "ACS Can Drive API running",
            "data_counts": {
                "students": student_count,
                "teachers": teacher_count,
                "donations": donation_count,
                "reservations": reservation_count
            }
        }
    except Exception as e:
        return {"status": "unhealthy", "database": "disconnected", "error": str(e)}

@app.post("/api/fix-homeroom-numbers")
def fix_homeroom_numbers_endpoint():
    """Fix homeroom numbers from float to proper string format"""
    from models import Student, Teacher
    
    try:
        db = get_db_simple()
        
        # Fix student homeroom numbers
        students = db.query(Student).all()
        student_fixes = 0
        
        for student in students:
            if student.homeroom_number:
                # Convert to string and clean up
                original = student.homeroom_number
                fixed = str(original).strip()
                
                # Remove .0 suffix if present
                if fixed.endswith('.0'):
                    fixed = fixed[:-2]
                
                # Pad with leading zeros if it's a number (e.g., 18 -> 018)
                if fixed.isdigit():
                    # Only pad if it's 1-2 digits (don't pad 118 -> 0118)
                    if len(fixed) <= 2:
                        fixed = fixed.zfill(3)  # Pad to 3 digits: 18 -> 018
                
                if str(original) != fixed:
                    student.homeroom_number = fixed
                    student_fixes += 1
        
        # Fix teacher homeroom numbers
        teachers = db.query(Teacher).all()
        teacher_fixes = 0
        
        for teacher in teachers:
            if teacher.homeroom_number:
                # Convert to string and clean up
                original = teacher.homeroom_number
                fixed = str(original).strip()
                
                # Remove .0 suffix if present
                if fixed.endswith('.0'):
                    fixed = fixed[:-2]
                
                # Pad with leading zeros if it's a number (e.g., 18 -> 018)
                if fixed.isdigit():
                    # Only pad if it's 1-2 digits (don't pad 118 -> 0118)
                    if len(fixed) <= 2:
                        fixed = fixed.zfill(3)  # Pad to 3 digits: 18 -> 018
                
                if str(original) != fixed:
                    teacher.homeroom_number = fixed
                    teacher_fixes += 1
        
        # Commit all changes
        db.commit()
        
        return {
            "success": True,
            "message": "Homeroom numbers fixed successfully",
            "student_fixes": student_fixes,
            "teacher_fixes": teacher_fixes,
            "total_fixes": student_fixes + teacher_fixes
        }
        
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.options("/{path:path}")
def options_handler(path: str):
    return {"message": "OK"}

# =============================================================================
# STUDENT MANAGEMENT ENDPOINTS
# =============================================================================



@app.get("/api/events/{event_id}/students/search")
def search_students_direct(event_id: int, q: str):
    from database import get_db
    from models import Student
    try:
        db = get_db_simple()
        query = db.query(Student).filter(Student.event_id == event_id)
        
        # Search by name (first_name + last_name)
        if q:
            query = query.filter((Student.first_name + " " + Student.last_name).ilike(f"%{q}%"))
        
        students = query.limit(10).all()  # Limit to 10 results for autocomplete
        return [
            {
                "id": s.id,
                "name": (f"{(s.first_name or '').strip()} {(s.last_name or '').strip()}".strip()) or None,
                "first_name": s.first_name,
                "last_name": s.last_name,
                "grade": s.grade,
                "homeroomNumber": s.homeroom_number,
                "homeroomTeacher": s.homeroom_teacher,
                "totalCans": s.total_cans or 0,
            }
            for s in students
        ]
    except Exception as e:
        return {"error": str(e)}

@app.post("/api/events/{event_id}/students/verify")
def verify_student_direct(event_id: int, payload: dict):
    from database import get_db
    from models import Student
    try:
        db = get_db_simple()
        
        # Accept flexible payload shapes from frontend
        name = (payload.get('name') or '').strip()
        first_name = (payload.get('first_name') or '').strip()
        last_name = (payload.get('last_name') or '').strip()
        grade = payload.get('grade')
        homeroom_number = payload.get('homeroom_number')
        homeroom_teacher = payload.get('homeroom_teacher')
        
        if name and not (first_name or last_name):
            # Parse "Last, First" or "First Last"
            if ',' in name:
                parts = [p.strip() for p in name.split(',', 1)]
                if len(parts) == 2:
                    last_name, first_name = parts[0], parts[1]
            else:
                parts = [p.strip() for p in name.split(' ') if p.strip()]
                if len(parts) >= 2:
                    first_name = parts[0]
                    last_name = ' '.join(parts[1:])
        
        # Search for student
        query = db.query(Student).filter(Student.event_id == event_id)
        
        if first_name and last_name:
            query = query.filter(
                Student.first_name.ilike(f"%{first_name}%"),
                Student.last_name.ilike(f"%{last_name}%")
            )
        elif name:
            query = query.filter((Student.first_name + " " + Student.last_name).ilike(f"%{name}%"))
        
        if grade:
            query = query.filter(Student.grade == float(grade))
        if homeroom_number:
            query = query.filter(Student.homeroom_number.ilike(f"%{homeroom_number}%"))
        if homeroom_teacher:
            query = query.filter(Student.homeroom_teacher.ilike(f"%{homeroom_teacher}%"))
        
        student = query.first()
        
        if student:
            return {
                "id": student.id,
                "name": f"{student.first_name} {student.last_name}",
                "first_name": student.first_name,
                "last_name": student.last_name,
                "grade": student.grade,
                "homeroom_number": student.homeroom_number,
                "homeroom_teacher": student.homeroom_teacher,
                "total_cans": student.total_cans or 0,
            }
        else:
            return {"error": "Student not found in roster"}
            
    except Exception as e:
        return {"error": str(e)}


# =============================================================================
# LEADERBOARD ENDPOINTS
# =============================================================================

@app.get("/api/events/{event_id}/leaderboard")
def get_leaderboard(event_id: int):
    """
    Get comprehensive leaderboard data for a specific event
    
    This endpoint calculates and returns leaderboard data including:
    - Top 50 students ranked by total cans collected
    - Top 50 teachers ranked by total cans collected
    - Top 50 classes (homerooms) ranked by combined student + teacher cans
    - Top 50 grades ranked by total cans from students in that grade
    - Class buyout eligibility data (classes that have collected 10x their student count)
    - Total cans collected across the entire event
    
    The leaderboard uses a dual calculation method:
    1. Sums all donations from the Donation table (source of truth)
    2. Sums all total_cans fields from Student and Teacher tables (respects manual edits)
    3. Uses the maximum of both to ensure accuracy
    
    Returns:
        dict: JSON object containing all leaderboard categories and totals
    """
    from models import Student, Teacher, Donation
    from collections import defaultdict
    from sqlalchemy import text
    
    try:
        # Get database connection - using simple connection for SQLite
        db = get_db_simple()
        
        # Query all students and teachers for the specified event
        # These are the primary data sources for leaderboard calculations
        students = db.query(Student).filter(Student.event_id == event_id).all()
        teachers = db.query(Teacher).filter(Teacher.event_id == event_id).all()
        
        if not students and not teachers:
            return {
                "topStudents": [],
                "topClasses": [],
                "topGrades": [],
                "topTeachers": [],
                "classBuyout": [],
                "allClassBuyout": [],
                "totalCans": 0
            }
        
        # Calculate total cans using dual method for accuracy
        # Method 1: Sum all donations from Donation table (source of truth for transactions)
        donations = db.query(Donation).filter(Donation.event_id == event_id).all()
        total_cans_from_donations = sum(donation.amount or 0 for donation in donations)
        
        # Method 2: Sum total_cans fields from Student and Teacher tables
        # This respects manual edits made by admins in the admin panel
        student_cans = sum(student.total_cans or 0 for student in students)
        teacher_cans = sum(teacher.total_cans or 0 for teacher in teachers)
        total_cans_from_fields = student_cans + teacher_cans
        
        # Use the maximum of both methods to ensure we capture all cans
        # This handles cases where:
        # - Donations were recorded but not yet reflected in total_cans fields
        # - Manual edits were made to total_cans fields
        total_cans = max(total_cans_from_donations, total_cans_from_fields)
        
        # Top Students - sort by total_cans descending
        students_sorted = sorted(students, key=lambda s: s.total_cans or 0, reverse=True)
        top_students = []
        for i, student in enumerate(students_sorted[:50]):
            top_students.append({
                "rank": i + 1,
                "name": f"{student.first_name} {student.last_name}".strip(),
                "grade": student.grade,
                "homeroomNumber": student.homeroom_number,
                "totalCans": student.total_cans or 0
            })
        
        # Top Teachers - sort by total_cans descending
        teachers_sorted = sorted(teachers, key=lambda t: t.total_cans or 0, reverse=True)
        top_teachers = []
        for i, teacher in enumerate(teachers_sorted[:50]):
            top_teachers.append({
                "rank": i + 1,
                "name": teacher.full_name or f"{teacher.first_name} {teacher.last_name}".strip(),
                "homeroomNumber": teacher.homeroom_number,
                "totalCans": teacher.total_cans or 0
            })
        
        # Top Classes - group by teacher and homeroom (students + teachers with homerooms)
        class_groups = defaultdict(int)
        
        # Add student contributions to their classes
        for student in students:
            if student.homeroom_teacher and student.homeroom_number:
                key = f"{student.homeroom_teacher} {student.homeroom_number}".strip()
                class_groups[key] += student.total_cans or 0
        
        # Add teacher contributions to their homerooms
        for teacher in teachers:
            if teacher.homeroom_number:
                # Try to match with existing class keys first
                teacher_name = teacher.full_name or f"{teacher.first_name} {teacher.last_name}".strip()
                room = teacher.homeroom_number
                
                # Look for existing class with same room number
                found_existing = False
                for existing_key in class_groups.keys():
                    if room in existing_key:
                        class_groups[existing_key] += teacher.total_cans or 0
                        found_existing = True
                        break
                
                # If no existing class found, create new one
                if not found_existing:
                    key = f"{teacher_name} {room}".strip()
                    class_groups[key] += teacher.total_cans or 0
        
        classes_sorted = sorted(class_groups.items(), key=lambda x: x[1], reverse=True)
        top_classes = []
        for i, (class_name, cans) in enumerate(classes_sorted[:50]):
            # Split class name back into teacher and room
            parts = class_name.split(' ', 1)
            teacher = parts[0] if parts else ""
            room = parts[1] if len(parts) > 1 else ""
            
            top_classes.append({
                "rank": i + 1,
                "name": class_name,
                "homeroomNumber": room,
                "totalCans": cans
            })
        
        # Top Grades - group by grade (students only)
        grade_groups = defaultdict(int)
        for student in students:
            grade = str(student.grade or '').strip()
            if grade:  # Only count non-empty grades
                # Normalize grade to prevent duplicates (e.g., "12", "12.0", "12 " all become "12")
                try:
                    # Try to convert to int first, then back to string to normalize
                    normalized_grade = str(int(float(grade)))
                except (ValueError, TypeError):
                    # If conversion fails, use the original string
                    normalized_grade = grade
                grade_groups[normalized_grade] += student.total_cans or 0
        
        grades_sorted = sorted(grade_groups.items(), key=lambda x: x[1], reverse=True)
        top_grades = []
        for i, (grade, cans) in enumerate(grades_sorted[:50]):
            top_grades.append({
                "rank": i + 1,
                "grade": int(grade) if grade.isdigit() else grade,
                "totalCans": cans
            })
        
        # Class Buyout - calculate eligibility for each class
        class_buyout_data = []
        class_student_counts = defaultdict(int)
        class_can_totals = defaultdict(int)
        
        # Count students and cans per class
        for student in students:
            if student.homeroom_teacher and student.homeroom_number:
                class_key = f"{student.homeroom_teacher} {student.homeroom_number}".strip()
                class_student_counts[class_key] += 1
                class_can_totals[class_key] += student.total_cans or 0
        
        # Add teacher contributions to class buyout totals
        # This ensures consistency with the main Top Classes leaderboard
        for teacher in teachers:
            if teacher.homeroom_number:
                teacher_name = teacher.full_name or f"{teacher.first_name} {teacher.last_name}".strip()
                room = teacher.homeroom_number
                
                # Try to find existing class key to add to
                found_match = False
                for class_key in class_can_totals.keys():
                    # Match if room number is in the class key
                    if room in class_key:
                        class_can_totals[class_key] += teacher.total_cans or 0
                        found_match = True
                        break
                
                # If no match found (e.g. class has no students but has a teacher?), 
                # we generally don't create a new buyout entry since student_count would be 0
                # and thus filtered out later anyway.
        
        # Calculate buyout eligibility
        for class_name, student_count in class_student_counts.items():
            if student_count > 0:  # Exclude classes with 0 students
                required_cans = student_count * 10
                actual_cans = class_can_totals[class_name]
                is_eligible = actual_cans >= required_cans
                
                class_buyout_data.append({
                    "class_name": class_name,
                    "homeroom_teacher": class_name.split(' ', 1)[0] if ' ' in class_name else class_name,
                    "homeroom_number": class_name.split(' ', 1)[1] if ' ' in class_name else "",
                    "student_count": student_count,
                    "required_cans": required_cans,
                    "actual_cans": actual_cans,
                    "is_eligible": is_eligible,
                    "progress_percentage": min(100, (actual_cans / required_cans * 100)) if required_cans > 0 else 0
                })
        
        # Sort by eligibility (eligible first), then by actual cans (highest first)
        class_buyout_data.sort(key=lambda x: (not x["is_eligible"], -x["actual_cans"]))
        
        # Get first 20 eligible classes for leaderboard
        eligible_classes = [cls for cls in class_buyout_data if cls["is_eligible"]][:20]

        return {
            "topStudents": top_students,
            "topClasses": top_classes,
            "topGrades": top_grades,
            "topTeachers": top_teachers,
            "classBuyout": eligible_classes,
            "allClassBuyout": class_buyout_data,  # For admin panel
            "totalCans": total_cans
        }
        
    except Exception as e:
        print(f"Leaderboard error: {e}")
        return {
            "topStudents": [],
            "topClasses": [],
            "topGrades": [],
            "topTeachers": [],
            "classBuyout": [],
            "allClassBuyout": [],
            "totalCans": 0,
            "error": str(e)
        }
    finally:
        # Ensure database connection is closed
        try:
            db.close()
        except:
            pass


@app.get("/api/events/{event_id}/donations")
def list_donations_direct(event_id: int):
    from database import get_db
    from models import Donation
    try:
        db = get_db_simple()
        donations = db.query(Donation).filter(Donation.event_id == event_id).all()
        return donations
    except Exception as e:
        return {"error": str(e)}

@app.post("/api/events/{event_id}/donations")
def add_donation_direct(event_id: int, payload: dict):
    """
    Add a new donation for a student or teacher.
    
    This endpoint records a donation and updates the total_cans count for the
    respective student or teacher. Timestamps are stored in UTC after converting
    from Eastern Time for accurate daily tracking.
    
    Args:
        event_id: The event ID to add the donation to
        payload: JSON body with student_id or teacher_id, amount, and optional notes
    
    Returns:
        The created donation record
    """
    from database import get_db
    from models import Donation, Student, Teacher
    try:
        db = get_db_simple()
        
        # Create donation with Eastern timezone converted to UTC for storage
        from datetime import datetime, timezone, timedelta
        eastern_tz = timezone(timedelta(hours=-4))  # EDT (Eastern Daylight Time)
        now_eastern = datetime.now(eastern_tz)
        # Convert to UTC for database storage
        now_utc = now_eastern.astimezone(timezone.utc)
        
        donation = Donation(
            event_id=event_id,
            student_id=payload.get('student_id'),
            teacher_id=payload.get('teacher_id'),
            amount=payload.get('amount', 0),
            donation_date=now_utc
        )
        
        db.add(donation)
        
        # Update total_cans for student or teacher
        if payload.get('student_id'):
            student = db.query(Student).filter(
                Student.id == payload.get('student_id'), 
                Student.event_id == event_id
            ).first()
            
            if student:
                student.total_cans = (student.total_cans or 0) + payload.get('amount', 0)
            else:
                return {"error": "Student not found"}
                
        elif payload.get('teacher_id'):
            teacher = db.query(Teacher).filter(
                Teacher.id == payload.get('teacher_id'), 
                Teacher.event_id == event_id
            ).first()
            
            if teacher:
                teacher.total_cans = (teacher.total_cans or 0) + payload.get('amount', 0)
            else:
                return {"error": "Teacher not found"}
        else:
            return {"error": "Either student_id or teacher_id must be provided"}
        
        db.commit()
        db.refresh(donation)
        return donation
    except Exception as e:
        return {"error": str(e)}

# =============================================================================
# MAP RESERVATION ENDPOINTS
# =============================================================================

@app.get("/api/events/{event_id}/map-reservations")
def list_map_reservations_direct(event_id: int):
    """
    Get all map reservations for a specific event.
    
    Returns a list of all street reservations including geojson data for
    rendering street highlights on the map.
    
    Args:
        event_id: The event ID to get reservations for
    
    Returns:
        List of reservation objects with street and student information
    """
    from database import get_db
    from models import MapReservation
    try:
        db = get_db_simple()
        reservations = db.query(MapReservation).filter(MapReservation.event_id == event_id).all()
        return [
            {
                "id": r.id,
                "eventId": r.event_id,
                "studentId": r.student_id,
                "studentName": r.name,
                "streetName": r.street_name,
                "groupMembers": r.group_members or "",
                "geojson": r.geojson or "{}",  # Return the stored geojson data for map rendering
                "latitude": 0,  # Keep for backward compatibility
                "longitude": 0,
                "createdAt": r.timestamp.isoformat() if r.timestamp else None
            }
            for r in reservations
        ]
    except Exception as e:
        return {"error": str(e)}

@app.delete("/api/events/{event_id}/map-reservations/{reservation_id}")
def delete_map_reservation_direct(event_id: int, reservation_id: int):
    """
    Delete a map reservation.
    
    Args:
        event_id: The event ID
        reservation_id: The reservation ID to delete
    
    Returns:
        Success message or error
    """
    from models import MapReservation
    try:
        db = get_db_simple()
        
        # Find the reservation
        reservation = db.query(MapReservation).filter(
            MapReservation.id == reservation_id,
            MapReservation.event_id == event_id
        ).first()
        
        if not reservation:
            return {"error": "Reservation not found"}
        
        # Delete the reservation
        db.delete(reservation)
        db.commit()
        
        return {"message": "Reservation deleted successfully"}
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/events/{event_id}/map-reservations/export.csv")
def export_map_reservations_csv(event_id: int):
    """
    Export map reservations as a CSV file.
    
    Args:
        event_id: The event ID to export reservations for
    
    Returns:
        CSV file download with all reservation data
    """
    from database import get_db
    from models import MapReservation
    import csv
    import io
    
    try:
        db = get_db_simple()
        reservations = db.query(MapReservation).filter(MapReservation.event_id == event_id).all()
        
        # Create CSV content
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Street Name', 'Student Name', 'Group Members', 'Created At'])
        
        # Write data
        for r in reservations:
            writer.writerow([
                r.street_name or '',
                r.name or '',
                r.group_members or '',
                r.timestamp.isoformat() if r.timestamp else ''
            ])
        
        # Return CSV content
        csv_content = output.getvalue()
        output.close()
        
        from fastapi.responses import Response
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=map_reservations.csv"}
        )
        
    except Exception as e:
        print(f"CSV export error: {e}")
        from fastapi.responses import Response
        return Response(
            content=f"Error exporting CSV: {str(e)}",
            media_type="text/plain",
            status_code=500
        )

@app.post("/api/events/{event_id}/map-reservations")
def reserve_street_direct(event_id: int, payload: dict):
    """
    Reserve a street for can collection.
    
    This endpoint handles street reservations for both students and teachers.
    It checks for existing reservations to prevent conflicts and supports
    updating existing reservations by the same person.
    
    The geojson field stores the street path coordinates for map highlighting.
    
    Args:
        event_id: The event ID
        payload: JSON body with name, street_name, student_id, group_members, and geojson
    
    Returns:
        The created reservation object
    """
    from database import get_db
    from models import MapReservation
    try:
        db = get_db_simple()
        
        # Get the street names to check
        street_names = payload.get('street_name', '')
        if not street_names:
            return {"error": "No street name provided"}
        
        # Split comma-separated street names (supports multiple street selection)
        streets_to_check = [street.strip() for street in street_names.split(',')]
        
        # Check if any of these streets are already reserved by someone else
        for street in streets_to_check:
            # Extract just the street name (before the first comma) for comparison
            street_name_only = street.split(',')[0].strip() if ',' in street else street
            
            # Check for existing reservations with similar street names
            existing = db.query(MapReservation).filter(
                MapReservation.event_id == event_id,
                MapReservation.street_name.ilike(f'{street_name_only}%')
            ).first()
            
            if existing:
                # Check if it's the same person trying to reserve (allow editing)
                existing_name = existing.name or ''
                current_name = payload.get('name') or payload.get('student_name', '')
                
                if existing_name.lower() != current_name.lower():
                    # Different person trying to reserve - block it
                    if existing.student_id:
                        return {"error": f"Street '{street_name_only}' is already reserved by student: {existing_name}"}
                    else:
                        return {"error": f"Street '{street_name_only}' is already reserved by: {existing_name}"}
                else:
                    # Same person - delete existing reservation to allow update
                    db.delete(existing)
        
        # Create new reservation with geojson path data for map highlighting
        reservation = MapReservation(
            event_id=event_id,
            student_id=payload.get('student_id'),
            name=payload.get('name') or payload.get('student_name', 'Unknown'),
            street_name=payload.get('street_name'),
            group_members=payload.get('group_members', ''),
            geojson=payload.get('geojson', '{}')  # Store path coordinates for map rendering
        )
        
        db.add(reservation)
        db.commit()
        db.refresh(reservation)
        
        return {
            "id": reservation.id,
            "eventId": reservation.event_id,
            "studentId": reservation.student_id,
            "studentName": reservation.name,
            "streetName": reservation.street_name,
            "geojson": reservation.geojson or "{}",
            "latitude": 0,
            "longitude": 0,
            "createdAt": reservation.timestamp.isoformat() if reservation.timestamp else None
        }
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/events/{event_id}/students")
def get_students_direct(event_id: int, grade: str = None, homeroom: str = None, name: str = None, teacher: str = None):
    """
    Get all students for an event with optional filtering.
    
    Supports filtering by grade, homeroom number, student name, and teacher name.
    Used by the admin dashboard for student management and donation recording.
    
    Args:
        event_id: The event ID
        grade: Optional grade filter
        homeroom: Optional homeroom number filter
        name: Optional name search filter
        teacher: Optional teacher name filter
    
    Returns:
        List of student objects matching the filters
    """
    from database import get_db
    from models import Student
    from sqlalchemy import String
    try:
        db = get_db_simple()
        query = db.query(Student).filter(Student.event_id == event_id)
        
        # Apply optional filters
        if grade:
            query = query.filter(Student.grade == grade)
        if homeroom:
            query = query.filter(
                (Student.homeroom_number.ilike(f"%{homeroom}%")) |
                (Student.homeroom_number.cast(String).ilike(f"%{homeroom}%"))
            )
        if name:
            query = query.filter((Student.first_name + " " + Student.last_name).ilike(f"%{name}%"))
        if teacher:
            query = query.filter(Student.homeroom_teacher.ilike(f"%{teacher}%"))
        
        students = query.all()
        return [
            {
                "id": s.id,
                "name": (f"{(s.first_name or '').strip()} {(s.last_name or '').strip()}".strip()) or None,
                "first_name": s.first_name,
                "last_name": s.last_name,
                "grade": s.grade,
                "homeroomNumber": s.homeroom_number,
                "homeroomTeacher": s.homeroom_teacher,
                "totalCans": s.total_cans or 0,
            }
            for s in students
        ]
    except Exception as e:
        return {"error": str(e)}

@app.put("/api/events/{event_id}/students/{student_id}")
def update_student_direct(event_id: int, student_id: int, payload: dict):
    """
    Update a student's information.
    
    Allows updating student name, grade, homeroom, and total cans count.
    Used by admins for manual corrections and data management.
    
    Args:
        event_id: The event ID
        student_id: The student ID to update
        payload: JSON body with fields to update
    
    Returns:
        Updated student object
    """
    from models import Student
    try:
        db = get_db_simple()
        
        # Find the student
        student = db.query(Student).filter(
            Student.id == student_id,
            Student.event_id == event_id
        ).first()
        
        if not student:
            return {"error": "Student not found"}
        
        # Update fields if provided in payload
        if "totalCans" in payload:
            student.total_cans = payload["totalCans"]
        if "first_name" in payload:
            student.first_name = payload["first_name"]
        if "last_name" in payload:
            student.last_name = payload["last_name"]
        if "grade" in payload:
            student.grade = payload["grade"]
        if "homeroomNumber" in payload:
            student.homeroom_number = payload["homeroomNumber"]
        if "homeroomTeacher" in payload:
            student.homeroom_teacher = payload["homeroomTeacher"]
        
        db.commit()
        db.refresh(student)
        
        return {
            "id": student.id,
            "name": f"{student.first_name} {student.last_name}".strip(),
            "first_name": student.first_name,
            "last_name": student.last_name,
            "grade": student.grade,
            "homeroomNumber": student.homeroom_number,
            "homeroomTeacher": student.homeroom_teacher,
            "totalCans": student.total_cans or 0,
        }
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/events/{event_id}/teachers")
def get_teachers_direct(event_id: int):
    """
    Get all teachers for an event.
    
    Returns all teachers registered for the can drive event.
    Used by admin dashboard and donation recording.
    
    Args:
        event_id: The event ID
    
    Returns:
        List of teacher objects
    """
    from database import get_db
    from models import Teacher
    try:
        db = get_db_simple()
        teachers = db.query(Teacher).filter(Teacher.event_id == event_id).all()
        return [
            {
                "id": t.id,
                "first_name": t.first_name,
                "last_name": t.last_name,
                "full_name": t.full_name,
                "homeroom_number": t.homeroom_number,
                "total_cans": t.total_cans or 0,
            }
            for t in teachers
        ]
    except Exception as e:
        return {"error": str(e)}

# =============================================================================
# ROSTER UPLOAD ENDPOINTS
# =============================================================================

@app.post("/api/events/{event_id}/upload-roster")
async def upload_roster_direct(event_id: int, file: UploadFile = File(...)):
    """
    Upload student roster from Excel file.
    
    Parses an Excel file containing student information and imports them into
    the database. Supports various column formats and name patterns.
    
    Expected columns: Name (or First Name/Last Name), Grade, Homeroom, Teacher
    
    Args:
        event_id: The event ID to import students for
        file: Excel file (.xlsx) with student data
    
    Returns:
        Count of students added
    """
    from models import Student
    import openpyxl
    from io import BytesIO
    try:
        db = get_db_simple()
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        sheet = wb.active
        added = 0
        header_indexes = None
        
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            # Determine header mapping on first row
            if i == 0:
                headers = [str(c).strip().lower() if c is not None else '' for c in row]
                def find_any(keys):
                    return next((idx for idx, h in enumerate(headers) if any(k in h for k in keys)), None)
                header_indexes = {
                    'full_name': find_any(['name', 'student', 'full']),
                    'grade': find_any(['grade']),
                    'homeroom_number': (lambda x: x if x is not None else find_any(['room']))(find_any(['homeroom'])),
                    'homeroom_teacher': find_any(['teacher']),
                }
                # If headers not matched, fall back to first four columns
                if header_indexes['full_name'] is None and len(row) >= 4:
                    header_indexes = { 'full_name': 0, 'grade': 1, 'homeroom_number': 2, 'homeroom_teacher': 3 }
                continue

            def val(index):
                value = (str(row[index]).strip() if (index is not None and index < len(row) and row[index] is not None) else None)
                return value

            full_name = val(header_indexes['full_name'])
            grade = val(header_indexes['grade'])
            homeroom_number = val(header_indexes['homeroom_number'])
            homeroom_teacher = val(header_indexes['homeroom_teacher'])
            
            # Format homeroom number properly (normalize to 3 digits)
            if homeroom_number:
                if homeroom_number.endswith('.0'):
                    homeroom_number = homeroom_number[:-2]
                if homeroom_number.isdigit() and len(homeroom_number) <= 2:
                    homeroom_number = homeroom_number.zfill(3)

            # Parse full name into first and last name
            if not full_name:
                continue
            
            first_name = ''
            last_name = ''
            
            # Handle "Last, First" or "First Last" formats
            if ',' in full_name:
                parts = [p.strip() for p in full_name.split(',', 1)]
                if len(parts) == 2:
                    last_name, first_name = parts[0], parts[1]
            else:
                parts = [p.strip() for p in full_name.split(' ') if p.strip()]
                if len(parts) >= 2:
                    first_name = parts[0]
                    last_name = ' '.join(parts[1:])
                elif len(parts) == 1:
                    first_name = parts[0]
                    last_name = ''
            
            if not first_name:
                continue
                
            # Check for duplicate students
            existing = db.query(Student).filter(
                Student.event_id == event_id,
                Student.first_name == str(first_name).strip(),
                Student.last_name == str(last_name).strip(),
            ).first()
            
            if existing:
                continue
                
            student = Student(
                first_name=str(first_name).strip(),
                last_name=str(last_name).strip(),
                grade=str(grade).strip() if grade is not None else None,
                homeroom_number=str(homeroom_number).strip() if homeroom_number is not None else None,
                homeroom_teacher=str(homeroom_teacher).strip() if homeroom_teacher is not None else None,
                event_id=event_id,
            )
            db.add(student)
            added += 1
            
        db.commit()
        print(f"Student upload completed: Added {added} students from {file.filename}")
        return {"message": f"Added {added} students from {file.filename}", "added": added}
    except Exception as e:
        print(f"Student upload error: {e}")
        return {"error": str(e)}

@app.post("/api/events/{event_id}/upload-teachers")
async def upload_teachers_direct(event_id: int, file: UploadFile = File(...)):
    """
    Upload teacher roster from Excel file.
    
    Parses an Excel file containing teacher names and imports them into
    the database. Automatically matches teachers with student homeroom data.
    
    Args:
        event_id: The event ID to import teachers for
        file: Excel file (.xlsx) with teacher names
    
    Returns:
        Count of teachers added
    """
    from models import Teacher, Student
    import openpyxl
    from io import BytesIO
    try:
        db = get_db_simple()
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        sheet = wb.active
        added = 0
        
        # Get existing teachers from student data to match homeroom numbers
        existing_teachers = set()
        students = db.query(Student).filter(Student.event_id == event_id).all()
        for student in students:
            if student.homeroom_teacher:
                existing_teachers.add(student.homeroom_teacher.strip())
        
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:  # Skip header row
                continue
                
            teacher_name = str(row[0]).strip() if row[0] else None
            if not teacher_name:
                continue
            
            # Parse teacher name into first and last name
            if ',' in teacher_name:
                parts = [p.strip() for p in teacher_name.split(',', 1)]
                if len(parts) == 2:
                    last_name, first_name = parts[0], parts[1]
                else:
                    first_name = teacher_name
                    last_name = ""
            else:
                parts = [p.strip() for p in teacher_name.split(' ') if p.strip()]
                if len(parts) >= 2:
                    first_name = parts[0]
                    last_name = ' '.join(parts[1:])
                else:
                    first_name = teacher_name
                    last_name = ""
            
            # Check if teacher already exists
            existing_teacher = db.query(Teacher).filter(
                Teacher.event_id == event_id,
                Teacher.first_name == first_name,
                Teacher.last_name == last_name
            ).first()
            
            if existing_teacher:
                continue
            
            # Check if this teacher has a homeroom from student data
            homeroom_number = None
            for student in students:
                if student.homeroom_teacher and student.homeroom_teacher.strip() == teacher_name:
                    homeroom_number = student.homeroom_number
                    break
            
            # Format homeroom number properly (normalize to 3 digits)
            if homeroom_number:
                if str(homeroom_number).endswith('.0'):
                    homeroom_number = str(homeroom_number)[:-2]
                if str(homeroom_number).isdigit() and len(str(homeroom_number)) <= 2:
                    homeroom_number = str(homeroom_number).zfill(3)
            
            teacher = Teacher(
                first_name=first_name,
                last_name=last_name,
                full_name=teacher_name,
                event_id=event_id,
                homeroom_number=homeroom_number
            )
            db.add(teacher)
            added += 1
        
        db.commit()
        print(f"Teacher upload completed: Added {added} teachers from {file.filename}")
        return {"added": added, "message": f"Added {added} new teachers"}
    except Exception as e:
        print(f"Teacher upload error: {e}")
        return {"error": str(e)}

# =============================================================================
# DAILY DONORS ENDPOINT
# =============================================================================

@app.get("/api/events/{event_id}/daily-donors")
def get_daily_donors(event_id: int):
    """
    Get top donors for the current day.
    
    Returns today's top students, teachers, and grades based on donations
    made since 3 AM Eastern Time. Resets daily to show fresh daily rankings.
    
    The daily reset at 3 AM ensures that late-night donations are counted
    toward the previous day's totals.
    
    Args:
        event_id: The event ID
    
    Returns:
        Top students, teachers, and grades for the day with their daily cans
    """
    from models import Donation, Student, Teacher
    from datetime import datetime, date, timezone, timedelta
    from collections import defaultdict
    from sqlalchemy import text
    
    try:
        db = get_db_simple()
        
        # Test connection first
        db.execute(text("SELECT 1"))
        
        # Get today's date in Eastern Time (Windsor, Ontario)
        # Currently EDT (Eastern Daylight Time) = UTC-4
        eastern_tz = timezone(timedelta(hours=-4))
        now_eastern = datetime.now(eastern_tz)
        
        # Daily reset at 3 AM Eastern Time
        # If it's before 3 AM, use yesterday's date for the "day"
        if now_eastern.hour < 3:
            today = (now_eastern - timedelta(days=1)).date()
        else:
            today = now_eastern.date()
        
        # Convert Eastern time to UTC for database query
        # Day starts at 3 AM Eastern and ends at 2:59:59 AM Eastern the next day
        start_of_day_eastern = datetime.combine(today, datetime.min.time().replace(hour=3)).replace(tzinfo=eastern_tz)
        end_of_day_eastern = datetime.combine(today + timedelta(days=1), datetime.min.time().replace(hour=3)).replace(tzinfo=eastern_tz)
        
        # Convert to UTC for database comparison
        start_of_day_utc = start_of_day_eastern.astimezone(timezone.utc)
        end_of_day_utc = end_of_day_eastern.astimezone(timezone.utc)
        
        today_donations = db.query(Donation).filter(
            Donation.event_id == event_id,
            Donation.donation_date >= start_of_day_utc,
            Donation.donation_date < end_of_day_utc
        ).all()
        
        
        # Calculate ONLY today's donations (not cumulative totals)
        student_daily = defaultdict(int)
        teacher_daily = defaultdict(int)
        grade_daily = defaultdict(int)
        
        # Sum up only today's donations by student/teacher
        for donation in today_donations:
            if donation.student_id:
                student_daily[donation.student_id] += donation.amount or 0
            elif donation.teacher_id:
                teacher_daily[donation.teacher_id] += donation.amount or 0
        
        # Calculate grade totals from today's donations only
        if student_daily:
            students_with_donations = db.query(Student).filter(
                Student.id.in_(student_daily.keys()),
                Student.event_id == event_id
            ).all()
            
            for student in students_with_donations:
                daily_amount = student_daily[student.id]
                grade = str(student.grade or '').strip()
                if grade:
                    grade_daily[grade] += daily_amount
        
        
        # Get top 10 students for today
        top_students = []
        if student_daily:
            students = db.query(Student).filter(
                Student.id.in_(student_daily.keys()),
                Student.event_id == event_id
            ).all()
            
            student_rankings = []
            for student in students:
                daily_amount = student_daily[student.id]
                student_rankings.append({
                    "id": student.id,
                    "name": f"{student.first_name} {student.last_name}".strip(),
                    "dailyCans": daily_amount,
                    "grade": student.grade,
                    "homeroomNumber": student.homeroom_number
                })
            
            student_rankings.sort(key=lambda x: x["dailyCans"], reverse=True)
            top_students = student_rankings[:10]
            
            # Add rank
            for i, student in enumerate(top_students):
                student["rank"] = i + 1
        
        # Get top grades for today
        top_grades = []
        if grade_daily:
            grade_rankings = []
            for grade, daily_amount in grade_daily.items():
                grade_rankings.append({
                    "grade": grade,
                    "dailyCans": daily_amount
                })
            
            grade_rankings.sort(key=lambda x: x["dailyCans"], reverse=True)
            top_grades = grade_rankings[:10]
            
            # Add rank
            for i, grade in enumerate(top_grades):
                grade["rank"] = i + 1
        
        # Get top 10 teachers for today
        top_teachers = []
        if teacher_daily:
            teachers = db.query(Teacher).filter(
                Teacher.id.in_(teacher_daily.keys()),
                Teacher.event_id == event_id
            ).all()
            
            teacher_rankings = []
            for teacher in teachers:
                daily_amount = teacher_daily[teacher.id]
                teacher_rankings.append({
                    "id": teacher.id,
                    "name": teacher.full_name or f"{teacher.first_name} {teacher.last_name}".strip(),
                    "dailyCans": daily_amount,
                    "homeroomNumber": teacher.homeroom_number
                })
            
            teacher_rankings.sort(key=lambda x: x["dailyCans"], reverse=True)
            top_teachers = teacher_rankings[:10]
            
            # Add rank
            for i, teacher in enumerate(top_teachers):
                teacher["rank"] = i + 1
        
        return {
            "date": today.isoformat(),
            "topStudents": top_students,
            "topTeachers": top_teachers,
            "topGrades": top_grades
        }
        
    except Exception as e:
        print(f"Daily donors error: {e}")
        return {
            "date": date.today().isoformat(),
            "topStudents": [],
            "topTeachers": [],
            "topGrades": [],
            "error": str(e)
        }
    finally:
        # Ensure database connection is closed
        try:
            db.close()
        except:
            pass

# =============================================================================
# DATA MANAGEMENT ENDPOINTS
# =============================================================================

@app.delete("/api/events/{event_id}/reset")
def reset_event_direct(event_id: int, confirm: bool = False):
    """
    Reset all data for an event - DANGEROUS OPERATION.
    
    This endpoint deletes ALL students, donations, map reservations, and teachers
    for the specified event. Requires explicit confirmation via query parameter.
    
    Args:
        event_id: The event ID to reset
        confirm: Must be True to proceed with reset
    
    Returns:
        Counts of deleted records
    """
    from database import get_db
    from models import Student, Donation, MapReservation, Teacher
    try:
        if not confirm:
            return {"error": "Reset requires confirmation. Add ?confirm=true to the request."}
            
        db = get_db_simple()
        
        # Count existing data before deletion
        student_count = db.query(Student).filter(Student.event_id == event_id).count()
        donation_count = db.query(Donation).filter(Donation.event_id == event_id).count()
        reservation_count = db.query(MapReservation).filter(MapReservation.event_id == event_id).count()
        teacher_count = db.query(Teacher).filter(Teacher.event_id == event_id).count()
        
        print(f"RESET WARNING: Deleting {student_count} students, {donation_count} donations, {reservation_count} reservations, {teacher_count} teachers")
        
        # Delete all data for the event (order matters due to foreign keys)
        db.query(MapReservation).filter(MapReservation.event_id == event_id).delete()
        db.query(Donation).filter(Donation.event_id == event_id).delete()
        db.query(Student).filter(Student.event_id == event_id).delete()
        db.query(Teacher).filter(Teacher.event_id == event_id).delete()
        
        db.commit()
        return {
            "status": "ok", 
            "message": f"All data for event {event_id} has been reset",
            "deleted": {
                "students": student_count,
                "donations": donation_count,
                "reservations": reservation_count,
                "teachers": teacher_count
            }
        }
    except Exception as e:
        return {"error": str(e)}

@app.post("/api/events/{event_id}/map-reservations/import")
async def import_map_reservations_csv(event_id: int, file: UploadFile = File(...)):
    """
    Import map reservations from a CSV file.
    
    Restores previously exported reservations or bulk imports new ones.
    Expected columns: Student Name, Street Name, Group Members (optional)
    
    Args:
        event_id: The event ID to import reservations for
        file: CSV file with reservation data
    
    Returns:
        Count of added and skipped records
    """
    from models import MapReservation
    import csv
    import io
    import json
    
    try:
        print(f"Starting import of file: {file.filename}")
        db = get_db_simple()
        
        # Read CSV content
        content = await file.read()
        csv_content = content.decode('utf-8')
        
        csv_reader = csv.DictReader(io.StringIO(csv_content))
        
        added = 0
        skipped = 0
        for row_num, row in enumerate(csv_reader, 1):
            # Skip if required fields are missing
            if not row.get('Student Name') or not row.get('Street Name'):
                skipped += 1
                continue
                
            # Check if reservation already exists
            existing = db.query(MapReservation).filter(
                MapReservation.event_id == event_id,
                MapReservation.name == row['Student Name'],
                MapReservation.street_name == row['Street Name']
            ).first()
            
            if existing:
                skipped += 1
                continue
            
            # Create geojson with coordinates if available
            geojson_data = {}
            if row.get('Latitude') and row.get('Longitude'):
                try:
                    geojson_data = {
                        "lat": float(row['Latitude']),
                        "lng": float(row['Longitude']),
                        "group": row.get('Group Members', '')
                    }
                except (ValueError, TypeError):
                    geojson_data = {}
            
            # Create new reservation
            reservation = MapReservation(
                event_id=event_id,
                name=row['Student Name'],
                street_name=row['Street Name'],
                group_members=row.get('Group Members', ''),
                geojson=json.dumps(geojson_data) if geojson_data else '',
                student_id=None
            )
            
            db.add(reservation)
            added += 1
        
        db.commit()
        print(f"Import completed: {added} added, {skipped} skipped")
        return {"message": f"Successfully imported {added} reservations", "added": added, "skipped": skipped}
        
    except Exception as e:
        print(f"Import error: {str(e)}")
        return {"error": str(e)}

# =============================================================================
# EXPORT ENDPOINTS
# =============================================================================

@app.get("/api/events/{event_id}/leaderboard/csv")
def export_leaderboard_csv(event_id: int):
    """
    Export leaderboard data as CSV file.
    
    Exports all students and teachers who have donated cans, sorted by
    total cans in descending order. Useful for record keeping and awards.
    
    Args:
        event_id: The event ID to export leaderboard for
    
    Returns:
        CSV file download with leaderboard data
    """
    from models import Student, Teacher
    import csv
    import io
    
    try:
        db = get_db_simple()
        
        # Get all students and teachers
        students = db.query(Student).filter(Student.event_id == event_id).all()
        teachers = db.query(Teacher).filter(Teacher.event_id == event_id).all()
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header with all requested fields
        writer.writerow(['Name', 'Grade', 'Homeroom Number', 'Homeroom Teacher', 'Total Cans', 'Rank'])
        
        # Write student data - ONLY students who have donated cans
        student_rankings = []
        for student in students:
            if student.total_cans and student.total_cans > 0:  # Only include students with cans
                student_rankings.append({
                    'name': f"{student.first_name} {student.last_name}".strip(),
                    'grade': student.grade,
                    'homeroom_number': student.homeroom_number,
                    'homeroom_teacher': student.homeroom_teacher,
                    'total_cans': student.total_cans or 0
                })
        
        # Sort by total cans descending
        student_rankings.sort(key=lambda x: x['total_cans'], reverse=True)
        
        # Write student data to CSV
        for i, student in enumerate(student_rankings):
            writer.writerow([
                student['name'],
                student['grade'] or '',
                student['homeroom_number'] or '',
                student['homeroom_teacher'] or '',
                student['total_cans'],
                i + 1
            ])
        
        # Write teacher data - ONLY teachers who have donated cans
        teacher_rankings = []
        for teacher in teachers:
            if teacher.total_cans and teacher.total_cans > 0:  # Only include teachers with cans
                teacher_rankings.append({
                    'name': teacher.full_name or f"{teacher.first_name} {teacher.last_name}".strip(),
                    'homeroom_number': teacher.homeroom_number,
                    'total_cans': teacher.total_cans or 0
                })
        
        # Sort by total cans descending
        teacher_rankings.sort(key=lambda x: x['total_cans'], reverse=True)
        
        # Write teacher data to CSV
        for i, teacher in enumerate(teacher_rankings):
            writer.writerow([
                teacher['name'],
                'Teacher',  # Grade field for teachers
                teacher['homeroom_number'] or '',
                '',  # No homeroom teacher for teachers themselves
                teacher['total_cans'],
                i + 1
            ])
        
        # Get CSV content
        csv_content = output.getvalue()
        output.close()
        
        # Close database connection
        db.close()
        
        from fastapi.responses import Response
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=leaderboard_export.csv"}
        )
        
    except Exception as e:
        print(f"CSV export error: {e}")
        from fastapi.responses import Response
        return Response(
            content=f"Error exporting CSV: {str(e)}",
            media_type="text/plain",
            status_code=500
        )
    finally:
        # Ensure database connection is closed
        try:
            db.close()
        except:
            pass
